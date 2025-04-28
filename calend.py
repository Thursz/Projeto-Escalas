import sqlite3
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Tentar importar a biblioteca de feriados
try:
    import holidays
    br_holidays = holidays.Brazil(subdiv='TO')
except ImportError:
    br_holidays = {}

# Feriados municipais de Palmas-TO (dia/mês)
palmas_holidays = {
    "20/05": "Aniversário de Palmas",
    # adicione outros feriados municipais aqui...
}

# Estilos de planilha
header_fill   = PatternFill(start_color="ADD8E6", fill_type="solid")
holiday_fill  = PatternFill(start_color="FFF2CC", fill_type="solid")
bold_font     = Font(bold=True)
center_align  = Alignment(horizontal="center")
thin_border   = Border(left=Side('thin'), right=Side('thin'),
                       top=Side('thin'), bottom=Side('thin'))

# --- Banco de Dados ---
def init_db():
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS funcionarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('12x36','estagiario')),
            escala_dias TEXT CHECK(escala_dias IN ('pares','ímpares')),
            turno TEXT CHECK(turno IN ('12h','12h noturno','6h'))
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS escalas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            turno TEXT NOT NULL,
            original BOOLEAN DEFAULT 1,
            FOREIGN KEY(funcionario_id) REFERENCES funcionarios(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS trocas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            funcionario_original INTEGER NOT NULL,
            funcionario_substituto INTEGER NOT NULL,
            FOREIGN KEY(funcionario_original) REFERENCES funcionarios(id),
            FOREIGN KEY(funcionario_substituto) REFERENCES funcionarios(id)
        )
    """)
    conn.commit()
    conn.close()

init_db()

# --- Funções de negócio ---

def listar_funcionarios():
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("SELECT id, nome, tipo FROM funcionarios ORDER BY nome")
    funcs = c.fetchall()
    conn.close()
    return funcs

def listar_para_escala():
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("SELECT id, tipo, escala_dias, turno FROM funcionarios")
    rows = c.fetchall()
    conn.close()
    return rows

def cadastrar_funcionario(nome, tipo, escala_dias, turno, mes_ano):
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("""
        INSERT INTO funcionarios (nome, tipo, escala_dias, turno)
        VALUES (?, ?, ?, ?)
    """, (nome, tipo, escala_dias, turno))
    fid = c.lastrowid
    conn.commit()
    conn.close()
    gerar_escala(fid, tipo, escala_dias, turno, mes_ano)

def gerar_escala(funcionario_id, tipo, escala_dias, turno, mes_ano):
    if mes_ano:
        try:
            mes, ano = map(int, mes_ano.split("/"))
        except:
            return
    else:
        hoje = datetime.now()
        mes, ano = hoje.month, hoje.year
    primeiro = datetime(ano, mes, 1)
    _, dias_no_mes = calendar.monthrange(ano, mes)

    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    for i in range(dias_no_mes):
        dia = primeiro + timedelta(days=i)
        if tipo == "12x36":
            if (escala_dias == "pares" and dia.day % 2 == 0) or \
               (escala_dias == "ímpares" and dia.day % 2 != 0):
                c.execute("""
                    INSERT OR IGNORE INTO escalas (funcionario_id, data, turno, original)
                    VALUES (?, ?, ?, 1)
                """, (funcionario_id, dia.strftime("%d/%m/%Y"), turno))
        else:  # estagiário
            if dia.weekday() < 5:  # segunda a sexta
                c.execute("""
                    INSERT OR IGNORE INTO escalas (funcionario_id, data, turno, original)
                    VALUES (?, ?, ?, 1)
                """, (funcionario_id, dia.strftime("%d/%m/%Y"), turno))
    conn.commit()
    conn.close()

def exportar_escalas_excel():
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    try:
        c.execute("SELECT DISTINCT turno FROM escalas ORDER BY turno")
        turnos = [r[0] for r in c.fetchall()]
        if not turnos:
            messagebox.showinfo("Info", "Não há escalas para exportar.")
            return

        wb = Workbook()
        wb.remove(wb.active)

        for turno in turnos:
            ws = wb.create_sheet(title=turno[:31])
            headers = ["Funcionário", "Turno", "Data", "Feriado", "Substituído Por"]
            ws.append(headers)
            # formata cabeçalho
            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.border = thin_border

            # busca escalas
            c.execute("""
                SELECT e.id, f.id, f.nome, e.data, e.original
                  FROM escalas e
                  JOIN funcionarios f ON e.funcionario_id = f.id
                 WHERE e.turno = ?
                 ORDER BY e.data, f.nome
            """, (turno,))
            for esc_id, fid, nome, data_str, orig_flag in c.fetchall():
                dt = datetime.strptime(data_str, "%d/%m/%Y").date()
                # identifica feriado
                feriado_nome = ""
                # nacional/estadual
                try:
                    feriado_nome = br_holidays.get(dt, "")
                except:
                    feriado_nome = br_holidays.get(dt, "")
                # municipal
                md = dt.strftime("%d/%m")
                if md in palmas_holidays:
                    if feriado_nome:
                        feriado_nome += f"; {palmas_holidays[md]}"
                    else:
                        feriado_nome = palmas_holidays[md]

                # substituição
                substituido_por = ""
                if orig_flag == 0:
                    c.execute("""
                        SELECT fs.nome
                          FROM trocas t
                          JOIN funcionarios fs ON t.funcionario_substituto = fs.id
                         WHERE t.data = ? AND t.funcionario_original = ?
                    """, (data_str, fid))
                    res = c.fetchone()
                    substituido_por = res[0] if res else ""
                else:
                    # pular registro do substituto
                    c.execute("""
                        SELECT 1 FROM trocas
                         WHERE data = ? AND funcionario_substituto = ?
                    """, (data_str, fid))
                    if c.fetchone():
                        continue

                # escreve linha
                ws.append([nome, turno, data_str, feriado_nome, substituido_por])
                # destaca linha de feriado
                if feriado_nome:
                    row = ws.max_row
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col_idx).fill = holiday_fill

            # formata dados e ajusta colunas
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar planilha"
        )
        if path:
            wb.save(path)
            messagebox.showinfo("Sucesso", "Planilha salva com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")
    finally:
        conn.close()

def realizar_troca():
    orig = funcionario_original_var.get()
    sub  = funcionario_substituto_var.get()
    data_txt = troca_data_entry.get().strip()
    if not orig or not sub:
        messagebox.showerror("Erro", "Selecione ambos os funcionários.")
        return
    try:
        id_orig = int(orig.split(" - ")[0])
        id_sub  = int(sub.split(" - ")[0])
    except:
        messagebox.showerror("Erro", "Formato inválido.")
        return
    if id_orig == id_sub:
        messagebox.showerror("Erro", "Original e substituto devem ser diferentes.")
        return
    try:
        dt = datetime.strptime(data_txt, "%d/%m/%Y")
    except:
        messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.")
        return
    dstr = dt.strftime("%d/%m/%Y")

    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("SELECT 1 FROM trocas WHERE funcionario_original = ? AND data = ?", (id_orig, dstr))
    if c.fetchone():
        messagebox.showerror("Erro", "Já foi substituído neste dia.")
        conn.close()
        return

    c.execute("SELECT id, original FROM escalas WHERE funcionario_id = ? AND data = ?", (id_orig, dstr))
    row = c.fetchone()
    if not row:
        messagebox.showerror("Erro", "Funcionário não estava escalado neste dia.")
        conn.close()
        return
    esc_id, flag = row
    if flag == 0:
        messagebox.showerror("Erro", "Escala já não é original.")
        conn.close()
        return

    c.execute("SELECT turno FROM funcionarios WHERE id = ?", (id_sub,))
    res = c.fetchone()
    if not res:
        messagebox.showerror("Erro", "Turno do substituto não encontrado.")
        conn.close()
        return
    turno_sub = res[0]

    try:
        conn.execute("BEGIN")
        c.execute("""
            INSERT INTO trocas (data, funcionario_original, funcionario_substituto)
            VALUES (?, ?, ?)
        """, (dstr, id_orig, id_sub))
        c.execute("UPDATE escalas SET original = 0 WHERE id = ?", (esc_id,))
        c.execute("""
            INSERT OR IGNORE INTO escalas (funcionario_id, data, turno, original)
            VALUES (?, ?, ?, 1)
        """, (id_sub, dstr, turno_sub))
        conn.commit()
        messagebox.showinfo("Sucesso", "Troca realizada com sucesso!")
        funcionario_original_var.set("")
        funcionario_substituto_var.set("")
        troca_data_entry.delete(0, tk.END)
        atualizar_lista()
    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Erro", f"Falha ao registrar troca:\n{e}")
    finally:
        conn.close()

def cadastrar():
    nome        = nome_entry.get().strip()
    tipo        = tipo_var.get()
    escala_dias = escala_dias_var.get() if tipo == "12x36" else None
    turno       = turno_var.get()
    mes_ano     = mes_ano_var.get().strip() or None

    if mes_ano:
        try:
            datetime.strptime(mes_ano, "%m/%Y")
        except:
            messagebox.showerror("Erro", "Mês/Ano inválido. Use MM/YYYY.")
            return
    if not nome or tipo not in ["12x36", "estagiario"]:
        messagebox.showerror("Erro", "Preencha todos os campos corretamente.")
        return
    if tipo == "12x36" and escala_dias not in ["pares", "ímpares"]:
        messagebox.showerror("Erro", "Selecione pares ou ímpares.")
        return
    if not turno:
        messagebox.showerror("Erro", "Selecione um turno.")
        return

    cadastrar_funcionario(nome, tipo, escala_dias, turno, mes_ano)
    atualizar_lista()
    messagebox.showinfo("Sucesso", f"{nome} cadastrado para {mes_ano or 'mês atual'}!")
    nome_entry.delete(0, tk.END)
    tipo_var.set("")
    escala_dias_var.set("")
    turno_var.set("")
    mes_ano_var.set("")

def remover_funcionario():
    sel = tree.selection()
    if not sel:
        messagebox.showerror("Erro", "Selecione um funcionário para remover.")
        return
    fid = tree.item(sel[0])['values'][0]
    nome = tree.item(sel[0])['values'][1]
    if not messagebox.askyesno("Confirmar", f"Remover {nome}?"):
        return
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("DELETE FROM funcionarios WHERE id = ?", (fid,))
    c.execute("DELETE FROM escalas WHERE funcionario_id = ?", (fid,))
    conn.commit()
    conn.close()
    atualizar_lista()

def atualizar_comboboxes_troca():
    vals = [f"{i} - {n}" for i, n, _ in listar_funcionarios()]
    funcionario_original_combo['values'] = vals
    funcionario_substituto_combo['values'] = vals

def atualizar_combobox_editar():
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("""
        SELECT t.id, fo.nome, fs.nome, t.data
          FROM trocas t
          JOIN funcionarios fo ON t.funcionario_original = fo.id
          JOIN funcionarios fs ON t.funcionario_substituto = fs.id
         ORDER BY t.data DESC
    """)
    items = [f"{tid} - {orig} → {sub} em {dt}" for tid, orig, sub, dt in c.fetchall()]
    conn.close()
    troca_editar_combo['values'] = items

def editar_troca():
    sel = troca_editar_var.get().split(" - ")[0]
    nova = nova_data_entry.get().strip()
    if not sel or not nova:
        messagebox.showerror("Erro", "Selecione a troca e informe a nova data.")
        return
    try:
        new_dt = datetime.strptime(nova, "%d/%m/%Y").strftime("%d/%m/%Y")
    except:
        messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.")
        return
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    c.execute("SELECT funcionario_original, funcionario_substituto, data FROM trocas WHERE id = ?", (sel,))
    row = c.fetchone()
    if not row:
        messagebox.showerror("Erro", "Troca não encontrada.")
        conn.close()
        return
    orig_id, sub_id, old_dt = row
    try:
        conn.execute("BEGIN")
        c.execute("UPDATE trocas SET data = ? WHERE id = ?", (new_dt, sel))
        c.execute("UPDATE escalas SET original = 1 WHERE funcionario_id = ? AND data = ?", (orig_id, old_dt))
        c.execute("DELETE FROM escalas WHERE funcionario_id = ? AND data = ?", (sub_id, old_dt))
        c.execute("UPDATE escalas SET original = 0 WHERE funcionario_id = ? AND data = ?", (orig_id, new_dt))
        c.execute("""
            INSERT OR IGNORE INTO escalas (funcionario_id, data, turno, original)
            VALUES (?, ?, (SELECT turno FROM funcionarios WHERE id = ?), 1)
        """, (sub_id, new_dt, sub_id))
        conn.commit()
        messagebox.showinfo("Sucesso", "Data de troca atualizada com sucesso!")
        nova_data_entry.delete(0, tk.END)
        atualizar_lista()
    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Erro", f"Falha ao editar troca:\n{e}")
    finally:
        conn.close()

def atualizar_escala():
    mes_ano = atualiza_mes_ano_var.get().strip()
    if not mes_ano:
        messagebox.showerror("Erro", "Informe Mês/Ano para atualizar.")
        return
    try:
        datetime.strptime(mes_ano, "%m/%Y")
    except:
        messagebox.showerror("Erro", "Formato de Mês/Ano inválido. Use MM/YYYY.")
        return
    conn = sqlite3.connect("escalas.db")
    c = conn.cursor()
    pattern = "%/" + mes_ano  # ex: "%/04/2025"
    c.execute("DELETE FROM escalas WHERE data LIKE ?", (pattern,))
    conn.commit()
    conn.close()
    for fid, tipo, esc_dias, turno in listar_para_escala():
        gerar_escala(fid, tipo, esc_dias, turno, mes_ano)
    messagebox.showinfo("Sucesso", f"Escala atualizada para {mes_ano}!")
    atualizar_lista()

def atualizar_lista():
    for item in tree.get_children():
        tree.delete(item)
    for i, n, t in listar_funcionarios():
        tree.insert("", "end", values=(i, n, t))
    atualizar_comboboxes_troca()
    atualizar_combobox_editar()

# --- Interface Gráfica ---
root = tk.Tk()
root.title("Sistema de Escalas")
root.geometry("600x900")

# Cadastro de Funcionário
frame_cad = tk.LabelFrame(root, text="Cadastro de Funcionário", padx=10, pady=10)
frame_cad.grid(row=0, column=0, padx=10, pady=5, sticky="ew")
tk.Label(frame_cad, text="Nome:").grid(row=0, column=0, sticky="w")
nome_entry = tk.Entry(frame_cad, width=30)
nome_entry.grid(row=0, column=1, pady=2)
tk.Label(frame_cad, text="Tipo:").grid(row=1, column=0, sticky="w")
tipo_var = tk.StringVar()
tk.Radiobutton(frame_cad, text="12x36", variable=tipo_var, value="12x36").grid(row=1, column=1, sticky="w")
tk.Radiobutton(frame_cad, text="Estagiário", variable=tipo_var, value="estagiario").grid(row=1, column=2, sticky="w")
tk.Label(frame_cad, text="Escala Dias:").grid(row=2, column=0, sticky="w")
escala_dias_var = tk.StringVar()
ttk.Combobox(frame_cad, textvariable=escala_dias_var, values=["pares","ímpares"], width=15).grid(row=2, column=1, pady=2, sticky="w")
tk.Label(frame_cad, text="Turno:").grid(row=3, column=0, sticky="w")
turno_var = tk.StringVar()
ttk.Combobox(frame_cad, textvariable=turno_var, values=["12h","12h noturno","6h"], width=15).grid(row=3, column=1, pady=2, sticky="w")
tk.Label(frame_cad, text="Mês/Ano (MM/YYYY):").grid(row=4, column=0, sticky="w")
mes_ano_var = tk.StringVar()
tk.Entry(frame_cad, textvariable=mes_ano_var, width=15).grid(row=4, column=1, pady=2, sticky="w")
tk.Button(frame_cad, text="Cadastrar", command=cadastrar).grid(row=5, column=0, columnspan=3, pady=5)

# Lista de Funcionários
frame_list = tk.LabelFrame(root, text="Funcionários Cadastrados", padx=10, pady=10)
frame_list.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
tree = ttk.Treeview(frame_list, columns=("ID","Nome","Tipo"), show="headings", height=8)
for col, w in [("ID",50),("Nome",200),("Tipo",100)]:
    tree.heading(col, text=col)
    tree.column(col, width=w, anchor="center")
tree.grid(row=0, column=0, sticky="nsew")
ttk.Scrollbar(frame_list, orient="vertical", command=tree.yview).grid(row=0, column=1, sticky="ns")
tree.configure(yscrollcommand=frame_list.children['!scrollbar'].set)
tk.Button(frame_list, text="Remover", command=remover_funcionario).grid(row=1, column=0, pady=5)

# Troca de Turno
frame_troca = tk.LabelFrame(root, text="Troca de Turno", padx=10, pady=10)
frame_troca.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
tk.Label(frame_troca, text="Original:").grid(row=0, column=0, sticky="w")
funcionario_original_var = tk.StringVar()
funcionario_original_combo = ttk.Combobox(frame_troca, textvariable=funcionario_original_var, width=30)
funcionario_original_combo.grid(row=0, column=1, pady=2, sticky="w")
tk.Label(frame_troca, text="Substituto:").grid(row=1, column=0, sticky="w")
funcionario_substituto_var = tk.StringVar()
funcionario_substituto_combo = ttk.Combobox(frame_troca, textvariable=funcionario_substituto_var, width=30)
funcionario_substituto_combo.grid(row=1, column=1, pady=2, sticky="w")
tk.Label(frame_troca, text="Data (DD/MM/AAAA):").grid(row=2, column=0, sticky="w")
troca_data_entry = tk.Entry(frame_troca, width=15)
troca_data_entry.grid(row=2, column=1, pady=2, sticky="w")
tk.Button(frame_troca, text="Realizar Troca", command=realizar_troca).grid(row=3, column=0, columnspan=2, pady=5)

# Editar Troca
frame_editar = tk.LabelFrame(root, text="Editar Data de Troca", padx=10, pady=10)
frame_editar.grid(row=3, column=0, padx=10, pady=5, sticky="ew")
tk.Label(frame_editar, text="Selecione Troca:").grid(row=0, column=0, sticky="w")
troca_editar_var = tk.StringVar()
troca_editar_combo = ttk.Combobox(frame_editar, textvariable=troca_editar_var, width=40)
troca_editar_combo.grid(row=0, column=1, pady=2, sticky="w")
tk.Label(frame_editar, text="Nova Data (DD/MM/AAAA):").grid(row=1, column=0, sticky="w")
nova_data_entry = tk.Entry(frame_editar, width=15)
nova_data_entry.grid(row=1, column=1, pady=2, sticky="w")
tk.Button(frame_editar, text="Salvar Alteração", command=editar_troca).grid(row=2, column=0, columnspan=2, pady=5)

# Atualizar Escala
frame_atualizar = tk.LabelFrame(root, text="Atualizar Escala", padx=10, pady=10)
frame_atualizar.grid(row=4, column=0, padx=10, pady=5, sticky="ew")
tk.Label(frame_atualizar, text="Mês/Ano (MM/YYYY):").grid(row=0, column=0, sticky="w")
atualiza_mes_ano_var = tk.StringVar()
tk.Entry(frame_atualizar, textvariable=atualiza_mes_ano_var, width=15).grid(row=0, column=1, pady=2, sticky="w")
tk.Button(frame_atualizar, text="Atualizar Escala", command=atualizar_escala).grid(row=1, column=0, columnspan=2, pady=5)

# Exportar para Excel
frame_export = tk.Frame(root)
frame_export.grid(row=5, column=0, pady=10)
tk.Button(frame_export, text="Exportar Escalas para Excel", command=exportar_escalas_excel).pack()

# Layout de redimensionamento
root.grid_rowconfigure(1, weight=1)
root.grid_columnconfigure(0, weight=1)
frame_list.grid_rowconfigure(0, weight=1)
frame_list.grid_columnconfigure(0, weight=1)

# Inicializa listas
atualizar_lista()

root.mainloop()
